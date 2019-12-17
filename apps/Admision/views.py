from io import BytesIO

#Vista genérica para mostrar resultados
from django.views.generic.base import TemplateView
#Workbook nos permite crear libros en excel
from openpyxl import Workbook
#Nos devuelve un objeto resultado, en este caso un archivo de excel
from django.http.response import HttpResponse


from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.platypus import Image,Table, Spacer, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
from django.http import HttpResponse, JsonResponse
from datetime import datetime , timedelta
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.decorators import api_view
from rest_framework import generics, status
from rest_framework.generics import get_object_or_404
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView
from rest_framework.filters import SearchFilter
from rest_framework.permissions import IsAuthenticated
from apps.Consultorio.models import Cita
from rest_framework.decorators import api_view, permission_classes
from .serializers import DistritoSerializer, ProvinciaSerializer, DepartamentoSerializer, HistoriaSerializer, HistoriaViewSerializer, DistritosxProvincia, ProvinciasxDepartamento
#, GrupSangSerializer
from .models import HorarioCab, HorarioDet, Provincia, Distrito, Departamento, Historia#, GrupSang
from apps.Consultorio.models import Triaje,Consulta, Cita
from apps.Administrador.models import Especialidad
import requests
from .pagination import SmallSetPagination
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from datetime import datetime
from datetime import date
x,y=A4
ca=""
contador=0
continuador=0
nuevo=0
reingreso=0
dy=0
can=""
    

# class vistaGrupoSang(ModelViewSet):
#     queryset = GrupSang.objects.all()
#     serializer_class = GrupSangSerializer

#Vista general de todos los  Distritos, Get Post Put Delete

def ultimaHistoria(request):
    
    nro = Historia.objects.all().order_by('numeroHistoria').last()
    if not nro:
        return JsonResponse({'status':'No hay Historias'})
    else:
        return JsonResponse({'NroHistoria': str(nro)})

class vistaDistrito(ModelViewSet):
    queryset = Distrito.objects.all()
    serializer_class = DistritoSerializer
    pagination_class = SmallSetPagination 
    permission_classes = [IsAuthenticated]

#Vista general de todos las provincias , Get Post Put Delete
class vistaProvincia(ModelViewSet):
    queryset = Provincia.objects.all()
    serializer_class = ProvinciaSerializer
    pagination_class = SmallSetPagination
    permission_classes = [IsAuthenticated]

#Vista general de todos las provincias , Get Post Put Delete
class vistaDepartamento(ModelViewSet):
    queryset = Departamento.objects.all()
    serializer_class = DepartamentoSerializer
    permission_classes = [IsAuthenticated]

#Vista para crear historias , Get Post Put Delete
class vistaCrearHistoria(ModelViewSet):
    queryset = Historia.objects.all()
    serializer_class = HistoriaSerializer
    pagination_class = SmallSetPagination
    filter_backends = [SearchFilter]
    search_fields = ["dni"]
    #search_fields = ["numeroHistoria"]
    permission_classes = [IsAuthenticated]

#Vista general de todos las historias clinicas, Get Post Put Delete
class vistaHistoria(ModelViewSet):
    queryset = Historia.objects.all().order_by("-id")
    serializer_class = HistoriaViewSerializer
    pagination_class = SmallSetPagination
    filter_backends = [SearchFilter]
    search_fields = ["dni"]
    #search_fields = ["numeroHistoria"]
    permission_classes = [IsAuthenticated]

#Busqueda por numero de historia , Serializer muestra todas las historias
class BuscarHistoria(generics.ListAPIView):

    serializer_class = HistoriaViewSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        nro = self.request.query_params.get('nro')
        return Historia.objects.filter(numeroHistoria=nro)

#Busqueda de las historias clinicas por  DNI , Serializer muestra todas las historias
class BuscarDNIH(generics.RetrieveUpdateDestroyAPIView):

    lookup_field = 'dni'
    serializer_class = HistoriaViewSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Historia.objects.all()

#Busqueda historia por nombre del paciente , Serializer muestra todas las historias
class BuscarNombreH(generics.ListAPIView):
    
    serializer_class = HistoriaViewSerializer
    pagination_class = SmallSetPagination
    permission_classes = [IsAuthenticated]
     
    def get_queryset(self):
        #id = self.kwargs['id']
        nombre = self.request.query_params.get('nom')
        qs = Historia.objects.filter(nombres__icontains=nombre)
        if qs.all().count()<1:
            qs = Historia.objects.filter(apellido_paterno__icontains=nombre)
        return qs

#Busqueda de distritos , Serializer muestra todos los distritos
class BuscarDistrito(generics.RetrieveUpdateDestroyAPIView):
    lookup_field = 'id'
    serializer_class = DistritosxProvincia
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Provincia.objects.all()

#Busqueda de provincias, Serializer muestra todas las provincias
class BuscarProvincia(generics.RetrieveUpdateDestroyAPIView):
    lookup_field = 'id'
    serializer_class = ProvinciasxDepartamento
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Departamento.objects.all()

#Busqueda de Distritos, Serializer muestra todas los Distritos
class BuscarDistritos(generics.ListAPIView):
    lookup_field = 'id'
    serializer_class = DistritoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        id = self.request.query_params.get('id')
        return Distrito.objects.filter(provincia__id=id)


#Busqueda de provincias, Serializer muestra todas las provincias
class BuscarProvincias(generics.ListAPIView):
    serializer_class = ProvinciaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        id = self.request.query_params.get('id')
        return Provincia.objects.filter(departamento__id=id)


def reniecDatos(request,dni):
    token = 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJlbWFpbCI6Imp2aWNlbnRlZy45NkBnbWFpbC5jb20ifQ.MyaKW0GJOlNqoLSYq5Vj0OIo-8oAew5OB3PT3vfZDjs'
    r = requests.get('http://dniruc.apisperu.com/api/v1/dni/' + dni + '?token=' + token)
    data = r.json()
    #print(data)
      
    return JsonResponse(data)
#JULIO VICENTE HAY CITAS EL DIA DE HOY
def haycitas(request,fecha):
    citas= Cita.objects.filter(fechaAtencion=fecha)
    if citas.count()>0:
        return JsonResponse({'hayCitas':'si'})
    else:
        return JsonResponse({'hayCitas':'no'})


def cancelarCitasFecha(request):
    #permission_classes = [IsAuthenticated]
    fecha = datetime.today()
    fechaInicio = fecha + timedelta(days=-3)
    fechaInicio = fechaInicio.strftime("%Y-%m-%d")
    fechaFin = fecha + timedelta(days=-1)
    fechaFin = fechaFin.strftime("%Y-%m-%d")
    qs = Cita.objects.filter(fechaAtencion__range=[fechaInicio,fechaFin])
    if qs.count()!=0:
        qs.update(estadoCita="Cancelado")
        return JsonResponse({'status':'done'})
    else:
        return JsonResponse({'status':'No existen Citas'})


#Realizado por Julio Vicente: Historial Clinico,contiene datos & Consultas ,con su triaje, del paciente , utiliza libreria reportlab
def HistoriaPDF(request,dni):
    global c,dy
    historia=Historia.objects.filter(numeroHistoria=dni)
    width,height =A4
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Historial_'+""+historia[0].numeroHistoria.__str__()+""+'.pdf'
    buffer = BytesIO()
    c = canvas.Canvas(buffer,pagesize=A4)

    #Cabecera__________________________________________
    c.setLineWidth(.3)
    c.setFont('Helvetica',20)
    c.drawString(185,750,'HISTORIA CLINICA')
    c.drawString(405, 765, 'N° HISTORIA')
    c.setFont('Helvetica',16)
    c.drawString(402,725,historia[0].numeroHistoria.__str__())
    c.drawImage("apps/Laboratorio/static/Unsa.png",60,700,width=85, height=110, mask='auto')
    c.line(40,695,550,695)
    c.line(40,820,550,820)
    c.line(40,695,40,820)
    c.line(165,695,165,820)
    c.line(395,695,395,820)
    c.line(550,695,550,820)
    c.line(395,750,550,750)
    #Cabecera_____________________________________________
    #nombre
    c.setFont('Helvetica', 16)
    c.drawString(60,650,'Apellidos y Nombres')
    c.setFont('Helvetica', 16)
    c.drawString(230,650,historia[0].nombres.__str__()+" "+historia[0].apellido_paterno.__str__()+" "+historia[0].apellido_materno.__str__())
    c.line(40,665,550,665)
    c.line(40,645,550,645)

    c.line(220,645,220,665)
    c.line(40,645,40,665)
    c.line(550,645,550,665)
    #nombre
    #fecha
    c.drawString(70,610,'Fecha de  Nacimiento')
    c.drawString(90,590,historia[0].fechaNac.__str__())

    c.drawString(340,610,'Lugar de Nacimiento')
    c.drawString(350,590,historia[0].lugarNac.__str__())

    c.line(40,625,550,625)
    c.line(40,605,550,605)
    c.line(40,585,550,585)

    c.line(40,585,40,625)
    c.line(265,585,265,625)
    c.line(550,585,550,625)

    #fecha
    #DNI
    c.drawString(125,550,'DNI')
    c.drawString(105,530,historia[0].dni.__str__())

    c.drawString(340,550,'Edad')
    c.drawString(350,530,str(historia[0].edad()))

    c.line(40,565,550,565)
    c.line(40,545,550,545)
    c.line(40,525,550,525)

    c.line(40,525,40,565)
    c.line(265,525,265,565)
    c.line(550,525,550,565)

    #Direccion-Distrito
    c.drawString(100,490,'Dirección')
    c.drawString(50,470,historia[0].direccion.__str__())

    c.drawString(350,490,'Distrito')
    c.drawString(335,470,historia[0].distrito.__str__())

    c.line(40,505,550,505)
    c.line(40,485,550,485)
    c.line(40,465,550,465)

    c.line(40,465,40,505)
    c.line(265,465,265,505)
    c.line(550,465,550,505)

    #Sexo EstadoCivil Profesión/Ocupación Teléfono

    c.drawString(50,430,'Sexo')
    c.drawString(50,410,historia[0].sexo.__str__())

    c.drawString(140,430,'Estado Civil')
    c.drawString(140,410,historia[0].estadoCivil.__str__())

    c.drawString(260,430,'Profesión/Ocupación ')
    c.drawString(260,410,historia[0].ocupacion.__str__())

    c.drawString(425,430,'Teléfono')
    c.drawString(425,410,historia[0].telefono.__str__())



    c.line(40,445,550,445)
    c.line(40,425,550,425)
    c.line(40,405,550,405)

    c.line(40,405,40,445)
    c.line(135,405,135,445)
    c.line(255,405,255,445)
    c.line(415,405,415,445)    
    c.line(550,405,550,445)

    #GRADO de institucion Procedencia
    c.drawString(70,370,'Grado de Instrucción')
    c.drawString(50,350,historia[0].gradoInstruccion.__str__())

    c.drawString(350,370,'Procedencia')
    c.drawString(335,350,historia[0].procedencia.__str__())

    c.line(40,385,550,385)
    c.line(40,365,550,365)
    c.line(40,345,550,345)

    c.line(40,345,40,385)
    c.line(265,345,265,385)
    c.line(550,345,550,385)

     #Edad y Fecha de Apertura
    c.drawString(45,310,'Fecha de Apertura')
    c.drawString(45,290,historia[0].fechaReg.__str__())

    c.line(40,325,550,325)
    c.line(40,305,550,305)
    c.line(40,285,550,285)

    c.line(40,285,40,325)
    c.line(550,285,550,325)
     # Close the PDF object cleanly.
    c.showPage()
    
    cita = Cita.objects.filter(numeroHistoria=historia[0].id)

    p = ParagraphStyle('test')
    p.textColor = 'black'
    p.borderColor = 'white'
    p.alignment = TA_CENTER
    p.borderWidth = 1
    p.fontSize = 9
    dy=800
    count=0
    if (cita!=None):
                
        for citas in cita:
            triaje=Triaje.objects.filter(cita=citas.id)
            if triaje.count()>0:
                consulta=Consulta.objects.filter(pk=triaje[0].id.__str__())#Filtro de consultas por el id de triaje
                      
            if  triaje.count()>0 and consulta.count()>0: 
                imprimirTriaje(p,triaje[0])
                imprimirConsulta(p,consulta[0])    
                dibujarBorde()
                count=count+1
            if count==4:
                c.showPage()
                dy=800

            dy=dy-23 

    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response
#JULIO DibujarBordes      
def dibujarBorde():
    global c,dy
    c.line(35,dy+174.5,560,dy+174.5)
    c.line(35,dy+151.5,560,dy+151.5)
    c.line(35,dy,560,dy)
    c.line(35,dy,35,dy+174.5)
    c.line(560,dy,560,dy+174.5)
    
                    
                            
                    

                  
#JULIO VICENTE COLOCA LOS ELEMENTOS DE  TRIAJE DE CITA PARA HISTORIA PDF
def imprimirTriaje(p,triaje):
    global dy,c
    p.alignment = TA_CENTER
    
    fecha=Paragraph("Fecha ",p) 
    fecha.wrapOn(c,130,90)
    fecha.drawOn(c, 40, dy)  
    fecha=Paragraph(str(triaje.fechaReg),p) 
    fecha.wrapOn(c,130,90)
    fecha.drawOn(c, 40, dy-11.5)  
    talla=Paragraph("Talla: " + str(triaje.talla),p) 
    talla.wrapOn(c,130,90)
    talla.drawOn(c, 170, dy) 
    peso=Paragraph("Peso: " + str(triaje.peso),p) 
    peso.wrapOn(c,130,90)
    peso.drawOn(c, 300, dy)  
    temperatura=Paragraph("T°:" + str(triaje.temperatura),p) 
    temperatura.wrapOn(c,130,90)
    temperatura.drawOn(c, 430, dy)      
    frecuenciaR=Paragraph("F.R:" + str(triaje.frecuenciaR),p) 
    frecuenciaR.wrapOn(c,130,90)
    frecuenciaR.drawOn(c, 170, dy-11.5)   
    frecuenciaC=Paragraph("F.C:" + str(triaje.frecuenciaC),p) 
    frecuenciaC.wrapOn(c,130,90)
    frecuenciaC.drawOn(c, 300, dy-11.5)   
    frecuenciaC=Paragraph("P.A:" + str(triaje.presionArt),p) 
    frecuenciaC.wrapOn(c,130,90)
    frecuenciaC.drawOn(c, 430, dy-11.5)   
    dy=dy-23

#JULIO VICENTE COLOCA LOS ELEMENTOS DE CONSULTA DEL TRIAJE DE LA CITA PARA HISTORIA PDF
def imprimirConsulta(p,consulta):
    global dy,c
    p.alignment = TA_JUSTIFY
    motivo=Paragraph(" MOTIVO CONSULTA: " + str(consulta.motivoConsulta),p) 
    motivo.wrapOn(c,520,180)
    motivo.drawOn(c, 40, dy-23)  
    dy=dy-34.5
    apetito=Paragraph(" APETITO: " + str(consulta.apetito),p) 
    apetito.wrapOn(c,172,180)
    apetito.drawOn(c, 40, dy-11.5)  
    orina=Paragraph(" ORINA: " + str(consulta.orina),p) 
    orina.wrapOn(c,172,180)
    orina.drawOn(c, 212, dy-11.5) 
    deposiciones=Paragraph(" DEPOSICIONES: " + str(consulta.deposiciones),p) 
    deposiciones.wrapOn(c,172,180)
    deposiciones.drawOn(c,384 , dy-11.5) 
    dy=dy-23   
    examenFisico=Paragraph(" EXAMEN FISICO: " + str(consulta.examenFisico),p) 
    examenFisico.wrapOn(c,258,180)
    examenFisico.drawOn(c, 40, dy-11.5)  
    dy=dy-23 
    diagnostico=Paragraph(" DIAGNOSTICO: " + str(consulta.diagnostico),p) 
    diagnostico.wrapOn(c,258,180)
    diagnostico.drawOn(c, 40, dy-34.5)     
    tratamiento=Paragraph(" TRATAMIENTO: " + str(consulta.tratamiento),p) 
    tratamiento.wrapOn(c,258,180)
    tratamiento.drawOn(c, 300, dy-34.5) 
    dy=dy-34.5
    proximaCita=Paragraph(" PROXIMA CITA: " + str(consulta.proximaCita),p) 
    proximaCita.wrapOn(c,258,180)
    proximaCita.drawOn(c, 40, dy-23) 
    dy=dy-23     




               

#JULIO VICENTE: MUESTRA EL REPORTE DE CITAS DEL DIA CON ESTADO ATENDIDO Y CANCELADO
def reporteDiarioCitas(request):
    global y,ca,contador,nuevo,reingreso,continuado
    especialidades = Especialidad.objects.all()
    fecha = datetime.today()
    fecha=fecha.strftime("%Y-%m-%d")
    width,height =A4
   
    citas= Cita.objects.filter(fechaAtencion=fecha).order_by("especialidad","medico")
    if citas.count()!=0:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=ReporteDiarioCitas_'+fecha+'.pdf'
        buffer = BytesIO()

        ca = canvas.Canvas(buffer,pagesize=A4)
        p = ParagraphStyle('test')
        p.textColor = 'black'
        p.borderColor = 'black'
        p.alignment = TA_CENTER
        p.borderWidth = 0.5
        p.fontSize = 10
        y=800
        crearCabeceraEspecialidad(str(citas[0].especialidad),fecha)
        crearCabeceraMedico(str(citas[0].medico.nombres+" "+citas[0].medico.apellido_paterno+" "+ citas[0].medico.apellido_materno))
        crearEncabezados(p)
        for i in range(citas.count()):
            
            imprimir(p,citas[i])
            if citas.count()!= i+1 :
                if str(citas[i].especialidad) != str(citas[i+1].especialidad):
                    if contador!=0 or nuevo!=0 or reingreso !=0 or continuador!=0:
                        y=y-10
                        contadorEstados(p)
                    y=y-10
                    crearCabeceraEspecialidad(str(citas[i+1].especialidad),fecha)
                    


                if str(citas[i].medico) != str(citas[i+1].medico):
                    if contador!=0 or nuevo!=0 or reingreso !=0 or continuador!=0:
                        y=y-10
                        contadorEstados(p)
                    y=y-5
                    crearCabeceraMedico(str(citas[i+1].medico.nombres+" "+citas[i+1].medico.apellido_paterno+" "+ citas[i+1].medico.apellido_materno))
                    crearEncabezados(p)
                
            if citas.count()-1== i :
                y=y-10
                contadorEstados(p)

        ca.save()
        pdf = buffer.getvalue()
        buffer.close()

        response.write(pdf)
        return response 
    else: 
        return JsonResponse({'status':'FAIL'})

#JULIO VICENTE: CREA LA TABLA DONDE ESTAN TODOS LOS ESTADOS 
def contadorEstados(p):
    global y,ca,contador,nuevo,continuador,reingreso
    p.borderColor = 'white'
    if (y<24):
        ca.showPage()
        y=800
    nuev = Paragraph("Nuevos",p)
    nuev.wrapOn(ca,103,90)
    nuev.drawOn(ca,40 , y)
    reingre = Paragraph("Reingreso",p)
    reingre.wrapOn(ca,103,90)
    reingre.drawOn(ca, 143, y)
    atendidos = Paragraph("Atendidos",p)
    atendidos.wrapOn(ca,103,90)
    atendidos.drawOn(ca, 246, y)
    aten = Paragraph("Continuador",p)
    aten.wrapOn(ca,103,90)
    aten.drawOn(ca, 349, y)
    conti = Paragraph("Atenciones",p)
    conti.wrapOn(ca,103,90)
    conti.drawOn(ca, 452, y)
    y=y-11.5
    nuev = Paragraph(str(nuevo),p)
    nuev.wrapOn(ca,103,90)
    nuev.drawOn(ca,40 , y)
    reingre = Paragraph(str(reingreso),p)
    reingre.wrapOn(ca,103,90)
    reingre.drawOn(ca, 143, y)
    atendidos = Paragraph(str(reingreso+nuevo),p)
    atendidos.wrapOn(ca,103,90)
    atendidos.drawOn(ca, 246, y)
    conti = Paragraph(str(continuador),p)
    conti.wrapOn(ca,103,90)
    conti.drawOn(ca, 349, y)
    aten = Paragraph(str(contador),p)
    aten.wrapOn(ca,103,90)
    aten.drawOn(ca, 452, y)
    y=y-11.5
    contador=0
    nuevo=0
    continuador=0
    reingreso=0
    p.borderColor = 'black'
#JULIO VICENTE: CREA EL TITULO CON LA ESPECIALIDAD Y LA FECHA 
def crearCabeceraEspecialidad(especialidad,fecha):
    global y,ca,contador
    #contador=0
    p1 = ParagraphStyle('test')
    p1.textColor = 'black'
    p1.borderColor = 'white'
    p1.alignment = TA_CENTER
    p1.borderWidth = 1
    p1.fontSize = 20
    if (y<41):
        ca.showPage()
        y=800
    especialidad = Paragraph(str(especialidad),p1)
    especialidad.wrapOn(ca,500,0)
    especialidad.drawOn(ca, 40, y)
    p1.fontSize = 12
    fecha = Paragraph("Fecha:"+" "+str(fecha),p1)
    fecha.wrapOn(ca,500,0)
    fecha.drawOn(ca,40, y-25)
    y=y-40

# JULIO VICENTE: COLOCA EL NOMBRE DEL MEDICO EN PDF
def crearCabeceraMedico(medico):   
    global y,ca,contador
    contador=0
    if (y<20):
        ca.showPage()
        y=800
        
     
    ca.setFont('Helvetica',14)
    ca.drawString(80,y,medico)
    ca.drawString(40,y,"Dr(a).")
    y=y-20   
# JULIO VICENTE: COLOCA LOS ATRIBUTOS DE LA CITA
def imprimir(p,cita):
    global y,ca,contador,nuevo,continuador,reingreso
    contador=contador+1
    if (y<12):
        ca.showPage()
        y=800
    
    cont = Paragraph(str(contador),p)
    cont.wrapOn(ca,15,90)
    cont.drawOn(ca, 40, y)
    historia = Paragraph(str(cita.numeroHistoria),p)
    historia.wrapOn(ca,80,90)
    historia.drawOn(ca, 55, y)
    nombre = Paragraph(str(cita.numeroHistoria.nombres)+" "+str(cita.numeroHistoria.apellido_paterno)+" "+str(cita.numeroHistoria.apellido_materno),p)
    nombre.wrapOn(ca,200,90)
    nombre.drawOn(ca, 135, y)
    recibo = Paragraph(str(cita.numeroRecibo),p)
    recibo.wrapOn(ca,70,90)
    recibo.drawOn(ca, 335, y)
   
    if cita.condicion=="N":
        nuevo=nuevo+1
    if cita.condicion=="R":
        reingreso=reingreso+1
    if cita.condicion=="C":
        continuador=continuador+1
    turno = Paragraph(str(cita.turno),p)
    turno.wrapOn(ca,70,90)
    turno.drawOn(ca, 405, y)
    condicion = Paragraph(str(cita.condicion),p)
    condicion.wrapOn(ca,80,90)
    condicion.drawOn(ca, 475, y)
    y=y-11.5


# JULIO VICENTE: CREAR EL ENCABEZADO DE LA TABLA 
def crearEncabezados(p):
    global y,ca,contador
    if (y<12):
        ca.showPage()
        y=800
    cont = Paragraph("N°",p)
    cont.wrapOn(ca,15,90)
    cont.drawOn(ca, 40, y)
    historia = Paragraph("N° HISTORIA",p)
    historia.wrapOn(ca,80,90)
    historia.drawOn(ca, 55, y)
    nombre = Paragraph("Apellidos y Nombres",p)
    nombre.wrapOn(ca,200,90)
    nombre.drawOn(ca, 135, y)
    recibo = Paragraph("N° Recibo",p)
    recibo.wrapOn(ca,70,90)
    recibo.drawOn(ca, 335, y)
    turno = Paragraph("Turno",p)
    turno.wrapOn(ca,70,90)
    turno.drawOn(ca, 405, y)
    condicion = Paragraph("Condicion",p)
    condicion.wrapOn(ca,80,90)
    condicion.drawOn(ca, 475, y)
    y=y-11.5


 
#JULIO VICENTE: MUESTRA EL REPORTE DE CITAS RANGO DE FECHA CON ESTADO ATENDIDO Y CANCELADO
def reporteCitasRangoFecha(request,fecha):
    global y,ca,contador,nuevo,reingreso,continuado
    
    especialidades = Especialidad.objects.all()    
    width,height =A4
   
    citas= Cita.objects.filter(fechaAtencion=fecha).order_by("especialidad","medico")
    if citas.count()!=0:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=ReporteDiario_'+fecha+'.pdf'
        buffer = BytesIO()

        ca = canvas.Canvas(buffer,pagesize=A4)
        p = ParagraphStyle('test')
        p.textColor = 'black'
        p.borderColor = 'black'
        p.alignment = TA_CENTER
        p.borderWidth = 0.5
        p.fontSize = 10
        y=800
        crearCabeceraEspecialidad(str(citas[0].especialidad),fecha)
        crearCabeceraMedico(str(citas[0].medico.nombres+" "+citas[0].medico.apellido_paterno+" "+ citas[0].medico.apellido_materno))
        crearEncabezados(p)
        for i in range(citas.count()):
            
            imprimir(p,citas[i])
            if citas.count()!= i+1 :
                if str(citas[i].especialidad) != str(citas[i+1].especialidad):
                    if contador!=0 or nuevo!=0 or reingreso !=0 or continuador!=0:
                        y=y-10
                        contadorEstados(p)
                    y=y-10
                    crearCabeceraEspecialidad(str(citas[i+1].especialidad),fecha)
                    


                if str(citas[i].medico) != str(citas[i+1].medico):
                    if contador!=0 or nuevo!=0 or reingreso !=0 or continuador!=0:
                        y=y-10
                        contadorEstados(p)
                    y=y-5
                    crearCabeceraMedico(str(citas[i+1].medico.nombres+" "+citas[i+1].medico.apellido_paterno+" "+ citas[i+1].medico.apellido_materno))
                    crearEncabezados(p)
                
            if citas.count()-1== i :
                y=y-10
                contadorEstados(p)

        ca.save()
        pdf = buffer.getvalue()
        buffer.close()

        response.write(pdf)
        return response 
    else: 
        return JsonResponse({'status':'FAIL'})



#generar excel con todas las historias clinicas
class ReportehistoriasExcel(TemplateView):

        #Usamos el método get para generar el archivo excel
        def get(self, request, *args, **kwargs):
                #Obtenemos todas las historias de nuestra base de datos
                historias = Historia.objects.all()

                provincias = Provincia.objects.all()
                #Creamos el libro de trabajo
                wb = Workbook()
                #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
                ws = wb.active
                #En la celda A1 ponemos el texto 'REPORTE DE historias'
                ws['A1'] = 'REPORTE DE HISTORIAS CLINICAS'
                #Juntamos las celdas desde la A1 hasta la U1, formando una sola celda
                ws.merge_cells('A1:U1')


                #Creamos los encabezados desde la celda B3 hasta la E3

                ws['A2'] = 'NRO DE HISTORIA'
                ws['B2'] = 'DNI'
                ws['C2'] = 'NOMBRES'
                ws['D2'] = 'APELLIDO_PATERNO'
                ws['E2'] = 'APELLIDO_MATERNO'
                ws['F2'] = 'DIRECCION'
                ws['G2'] = 'DISTRITO'
                ws['H2'] = 'CELULAR'
                ws['I2'] = 'TELEFONO'
                ws['J2'] = 'LUGAR DE NACIMIENTO'
                ws['K2'] = 'FECHA_DE_NACIMIENTO'
                ws['L2'] = 'EDAD'
                ws['M2'] = 'SEXO'
                ws['N2'] = 'GRADO DE INSTRUCCION'
                ws['O2'] = 'OCUPACION'
                ws['P2'] = 'ESTADO CIVIL'
                ws['Q2'] = 'NACIONALIDAD'
                ws['R2'] = 'PROVINCIA'
                ws['S2'] = 'DEPARTAMENTO'
                ws['T2'] = 'PROCEDENCIA'
                ws['U2'] = 'APERTURA'
            
                cont=3
                #Recorremos el conjunto de historias y vamos escribiendo cada uno de los datos en las celdas
                for historia in historias:
                    if historia.dni == None:
                        dni = "00000000"
                    else:
                        dni = historia.dni

                    if historia.nombres ==None:
                        nombres = ""
                    else :
                        nombres = historia.nombres

                    if historia.apellido_paterno ==None:
                        apaterno = ""
                    else :
                        apaterno = historia.apellido_paterno

                    if historia.apellido_materno ==None:
                        amaterno = ""
                    else :
                        amaterno = historia.apellido_materno

                    if historia.sexo ==None:
                        sexo = ""
                    else :
                        sexo = historia.sexo

                    if historia.fechaNac ==None:
                        fechaNac = ""
                    else :
                        fechaNac = historia.fechaNac


                    if historia.celular ==None:
                        celular = ""
                    else :
                        celular = historia.celular

                    if historia.telefono ==None:
                        telefono  = ""
                    else :
                        telefono  = historia.telefono 

                    if historia.estadoCivil ==None:
                        estadoCivil = ""
                    else :
                        estadoCivil = historia.estadoCivil

                    if historia.gradoInstruccion ==None:
                        gradoInstruccion = ""
                    else :
                        gradoInstruccion = historia.gradoInstruccion

                    if historia.ocupacion ==None:
                        ocupacion = ""
                    else :
                        ocupacion = historia.ocupacion

                    if historia.fechaReg == None:
                        fechaReg = ""
                    else :
                        fechaReg = historia.fechaReg

                    if historia.direccion == None:
                        direccion = ""
                    else :
                        direccion = historia.direccion

                    if historia.nacionalidad == None:
                        nacionalidad = ""
                    else :
                        nacionalidad = historia.nacionalidad

                    if historia.lugarNac == None:
                        lugarNac = ""
                    else :
                        lugarNac = historia.lugarNac

                    if historia.procedencia == None:
                        procedencia = ""
                    else :
                        procedencia = historia.procedencia

                    if historia.distrito == None:
                        distrito = ""
                    else :
                        distrito = historia.distrito.nombre

                    if historia.provincia == None:
                        provincia = ""
                    else :
                        provincia = historia.provincia.nombre 

                    if historia.departamento == None:
                        departamento = ""
                    else :
                        departamento = historia.departamento.nombre

                    if historia.fechaNac==None:
                        edad = 100
                    else:
                        edad = int((datetime.now().date() - historia.fechaNac).days / 365.25)

                    ws.cell(row=cont,column=1).value = historia.numeroHistoria
                    ws.cell(row=cont,column=2).value = dni
                    ws.cell(row=cont,column=3).value = nombres
                    ws.cell(row=cont,column=4).value = apaterno
                    ws.cell(row=cont,column=5).value = amaterno
                    ws.cell(row=cont,column=6).value = direccion
                    ws.cell(row=cont,column=7).value = distrito
                    ws.cell(row=cont,column=8).value = celular
                    ws.cell(row=cont,column=9).value = telefono
                    ws.cell(row=cont,column=10).value = lugarNac
                    ws.cell(row=cont,column=11).value = fechaNac
                    ws.cell(row=cont,column=12).value = edad
                    ws.cell(row=cont,column=13).value = sexo
                    ws.cell(row=cont,column=14).value = gradoInstruccion
                    ws.cell(row=cont,column=15).value = ocupacion
                    ws.cell(row=cont,column=16).value = estadoCivil
                    ws.cell(row=cont,column=17).value = nacionalidad
                    ws.cell(row=cont,column=18).value = provincia    
                    ws.cell(row=cont,column=19).value = departamento
                    ws.cell(row=cont,column=20).value = procedencia
                    ws.cell(row=cont,column=21).value = fechaReg

            
                    cont = cont + 1

                #Establecemos el nombre del archivo
                nombre_archivo ="Reportehistoriasclinicas.xlsx"
                #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
                response = HttpResponse(content_type="application/ms-excel")
                contenido = "attachment; filename={0}".format(nombre_archivo)
                response["Content-Disposition"] = contenido
                wb.save(response)
                return response

